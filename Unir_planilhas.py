Nome_planilha_treinos = input('Digite o nome da planilha que contem os treinos dos alunos: ')
Nome_planilha_aluno = input('Digite o nome da planilha do aluno que será adicionado: ')
nome_aluno = input('Digite o nome do aluno: ')












import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd
import os
from datetime import datetime

workbook_treinos_indv = openpyxl.load_workbook(Nome_planilha_aluno+'.xlsx')
Nomes_folhas_tr_indv = workbook_treinos_indv.sheetnames
Nomes_folhas_tr_indv.sort()
input_sheet = workbook_treinos_indv[Nomes_folhas_tr_indv[-1]]

workbook_todos_treinos = openpyxl.load_workbook(Nome_planilha_treinos+'.xlsx')
workbook_todos_treinos_backup = openpyxl.load_workbook(Nome_planilha_treinos+'.xlsx')
if nome_aluno in workbook_todos_treinos.sheetnames:
    workbook_todos_treinos.remove(workbook_todos_treinos[nome_aluno])
matriz_sheet = workbook_todos_treinos["Matriz"]
output_sheet = workbook_todos_treinos.copy_worksheet(matriz_sheet)
output_sheet.title = nome_aluno

FIRST_PROGRAM_COL = 7 # G
COLUMNS_OCCUPIED_BY_PROGRAM = 15
COLUMNS_BETWEEN_PROGRAMS = 1

Frequencia = int((input_sheet["C" + str(2)].value).split('X')[0]) # quantos dias por semana
max_columns = FIRST_PROGRAM_COL + COLUMNS_OCCUPIED_BY_PROGRAM * Frequencia + COLUMNS_BETWEEN_PROGRAMS * (Frequencia - 1)

# Determine the total columns in the sheet
total_columns = output_sheet.max_column

# Delete columns that are not needed
if total_columns > max_columns:
    output_sheet.delete_cols(max_columns + 1, total_columns - max_columns)

# Copiar todas as linhas da planilha de entrada para a planilha de saída
for row_index, row in enumerate(input_sheet.iter_rows(min_row=1, max_row=input_sheet.max_row, min_col=1, max_col=input_sheet.max_column), start=1):
    for col_index, cell in enumerate(row, start=1):
        output_sheet.cell(row=row_index, column=col_index, value=cell.value)

#datavalidation
exercicios = pd.read_excel("..\\data\\Constantes.xlsx", sheet_name="dExercicios")
const = exercicios[exercicios['Secao'] == 's&p']
lista_exercicios_dv = list(const['Exercicio'])

if 'Listas_exrc' in workbook_todos_treinos.sheetnames:
    del workbook_todos_treinos['Listas_exrc']
ws_listas = workbook_todos_treinos.create_sheet(title='Listas_exrc')

for idx, val in enumerate(lista_exercicios_dv, start=1):
    ws_listas[f'A{idx}'] = val
ws_listas.sheet_state = 'hidden'

for sheet_name in workbook_todos_treinos.sheetnames:
    if sheet_name not in ['listas_exrc,Matriz']:
        ws = workbook_todos_treinos[sheet_name]

        # Criar a validação de dados referenciando o intervalo de células na outra planilha
        formula = f"='Listas_exrc'!$A$1:$A${len(lista_exercicios_dv)}"
        dv = DataValidation(type='list', formula1=formula, allow_blank=True)

        # Adicionar a validação de dados à planilha
        ws.add_data_validation(dv)

        # Aplicar a validação de dados a todas as células na coluna B
        for letter in ['J','Z','AP','BF','BV','CL']:
            for row in range(23, 33):
                cell = ws[f'{letter}{row}']
                dv.add(cell)

now = datetime.now()
data_hora = now.strftime("%Y-%m-%d_%H-%M-%S")
nome_arquivo = f"backup_{data_hora}.xlsx"
caminho_arquivo = os.path.join('..\\backups', nome_arquivo)
workbook_todos_treinos_backup.save(caminho_arquivo)
workbook_todos_treinos.save(Nome_planilha_treinos+'.xlsx')
print('União completa')