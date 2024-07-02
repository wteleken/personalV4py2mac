import pandas as pd
import openpyxl
import random
import numpy as np


FIRST_PROGRAM_LINE = 12

nome_novo_treino =input('digite o nome da planilha de treino do aluno: ')
nome_novo_treino = nome_novo_treino + '.xlsx'
# move exercise
MOVE_EXERCISE_LINE = 4  # which means the first exercise is in line 4 + top_left_program_line
MOVE_EXERCISE_1_COL = 3  # which means the first exercise is in column 4 + top_left_program_col
MOVE_EXERCISE_2_COL = 9  # which means the second exercise is 6 columns to the right of the first exercise
MOVE_SER_REP_DELTA_COL = -1  # which means ser rep is 1 column to the left of the exercise
MOVE_MAX_LINES = 6

# snp exercise
SNP_EXERCISE_LINE = 11
SNP_EXERCISE_1_COL = 3
SNP_SERIE_DELTA_COL = 3
SNP_REP_DELTA_COL = 4
SNP_BLOCO_DELTA_COL = -1
SNP_PM_DELTA_COL = -2  # todos relativos ao exercise_col
SNP_MAX_LINES = 10

# esd exercise
ESD_EXERCISE_LINE = 24
ESD_EXERCISE_1_COL = 2

FIRST_PROGRAM_COL = 7 # G
COLUMNS_OCCUPIED_BY_PROGRAM = 15
COLUMNS_BETWEEN_PROGRAMS = 1

OUTPUT_VOLUME_ROWS = 5 #relative to output header rows

# coleta informações de exercícios
exercicios = pd.read_excel("..\\data\\Constantes.xlsx", sheet_name="dExercicios")
temas = pd.read_excel("..\\data\\Constantes.xlsx", sheet_name="dTemas")
workbook = openpyxl.load_workbook(nome_novo_treino)

INPUT_HEADER_ROWS = 2
OUTPUT_HEADER_ROWS = 2

#info_sheet = workbook["Infos"]
matriz_sheet = workbook["Matriz0"]

"""
    This section is to choose what program is going to be created
"""
Nomes_folhas_tr_indv = workbook.sheetnames
Nomes_folhas_tr_indv = [word for word in Nomes_folhas_tr_indv if not word.startswith('In')]
Nomes_folhas_tr_indv.sort()
ultimo_treino = Nomes_folhas_tr_indv[-1]
num_ultimo_treino = int(ultimo_treino[-1])
n_program = num_ultimo_treino + 1
# define how many programs have happened
#n_program = info_sheet["B1"].value

# define the input and output sheets
input_sheet = workbook["In Programa " + str(n_program)] # Input Programa n_program
output_sheet = workbook.copy_worksheet(matriz_sheet)
output_sheet.title = "Programa " + str(n_program)
# testing_output_sheet = workbook.create_sheet("Teste Programa " + str(n_program))


"""
    This section is to collect inputs from the user and constants about the training program once it's chosen
"""

rows_to_skip = INPUT_HEADER_ROWS

# get divisao
divisao = input_sheet["B" + str(rows_to_skip + 1)].value # como os exercicios se dividem ao longo dos dias

# get frequencia
Frequencia = input_sheet["B" + str(rows_to_skip + 2)].value # quantos dias por semana

# get tema
Tema = input_sheet["B" + str(rows_to_skip + 3)].value # quantidade de repeticoes e series

# get Maturidade
Maturidade = input_sheet["H" + str(rows_to_skip + 2)].value # quantos dias por semana

# Getting Volumes
Volume_DQ = input_sheet["B" + str(rows_to_skip + 5)].value
Volume_DJ = input_sheet["C" + str(rows_to_skip + 5)].value
Volume_EMP = input_sheet["D" + str(rows_to_skip + 5)].value
Volume_PUX = input_sheet["E" + str(rows_to_skip + 5)].value
Volume_ESTAB = input_sheet["F" + str(rows_to_skip + 5)].value
Volume_POT = input_sheet["G" + str(rows_to_skip + 5)].value
Volume_AUX = input_sheet["H" + str(rows_to_skip + 5)].value

# Printing Volumes
output_sheet.cell(row=OUTPUT_HEADER_ROWS,column=3,value=str(Frequencia)+'X / SEMANA')
output_sheet.cell(row=OUTPUT_HEADER_ROWS,column=5,value=divisao)
output_sheet.cell(row=OUTPUT_HEADER_ROWS+2,column=3,value=Tema)
output_sheet.cell(row=OUTPUT_HEADER_ROWS+OUTPUT_VOLUME_ROWS,column=3,value=Volume_DQ)
output_sheet.cell(row=OUTPUT_HEADER_ROWS+OUTPUT_VOLUME_ROWS,column=4,value=Volume_DJ)
output_sheet.cell(row=OUTPUT_HEADER_ROWS+OUTPUT_VOLUME_ROWS,column=5,value=Volume_EMP)
output_sheet.cell(row=OUTPUT_HEADER_ROWS+OUTPUT_VOLUME_ROWS,column=6,value=Volume_PUX)
output_sheet.cell(row=OUTPUT_HEADER_ROWS+OUTPUT_VOLUME_ROWS,column=7,value=Volume_ESTAB)
output_sheet.cell(row=OUTPUT_HEADER_ROWS+OUTPUT_VOLUME_ROWS,column=8,value=Volume_POT)
output_sheet.cell(row=OUTPUT_HEADER_ROWS+OUTPUT_VOLUME_ROWS,column=9,value=Volume_AUX)

# add volumes to the dictionary
volumes = {
    "DQ": Volume_DQ,
    "DJ": Volume_DJ,
    "EMP": Volume_EMP,
    "PUX": Volume_PUX,
    "ESTAB": Volume_ESTAB,
    "POT": Volume_POT,
    "AUX": Volume_AUX
}

print("Este é o programa numero", n_program)
print("Divisão:", divisao)
print("Frequencia:", Frequencia)
print("Tema:", Tema)
print("Volume:", volumes)
print("Iniciando a geração do programa...")

max_columns = FIRST_PROGRAM_COL + COLUMNS_OCCUPIED_BY_PROGRAM * Frequencia + COLUMNS_BETWEEN_PROGRAMS * (Frequencia - 1)

# Determine the total columns in the sheet
total_columns = output_sheet.max_column

# Delete columns that are not needed
if total_columns > max_columns:
    output_sheet.delete_cols(max_columns + 1, total_columns - max_columns)

Tema_escolhido = temas[temas["Tema"] == Tema]

def troca(lista,position, maxrep, minrep):
    if (all(minrep <= x <= maxrep for x in lista) == True) or lista[-1] == 0:
        if lista[-1] == 0:
            lista.remove(0)
        if (abs(position) <= len(lista)) and lista[position] == maxrep and (lista[-1] + 1 < maxrep):
            lista[-1] = lista[-1] + 1
            lista[position] = lista[position] - 1
        return lista
    else:
        if abs(position) > len(lista):
            position = -2
        lista[-1] = lista[-1] + 1
        lista[position] = lista[position] - 1
        troca(lista,position-1, maxrep, minrep)
    return lista

#{dicionario} com a correspondencia entre os tipos de exercicio DQ, DJ,... e uma [lista] com a quantidade
#de repetições para cada exercicio #ex: DQ = [4,3,3]
dict_lista_volumes = {}
for key,value in volumes.items():
    resto = value % Tema_escolhido['Max_Serie'].iloc[0]
    lista = [Tema_escolhido['Max_Serie'].iloc[0]] * (value // Tema_escolhido['Max_Serie'].iloc[0])
    lista.append(resto)
    lista = troca(lista, -2, Tema_escolhido['Max_Serie'].iloc[0], Tema_escolhido['Min_Serie'].iloc[0])
    dict_lista_volumes[key] = lista

exercicios = exercicios.dropna(subset=['Prioridade no dia'])
exercicios = exercicios[exercicios['Maturidade']<=Maturidade] #filtra por maturidade minima


# padronizacao do numero de repeticoes
n_reps = (str(Tema_escolhido["Min_Rep"].values[0]) + '-' + str(Tema_escolhido["Max_Rep"].values[0]))

# dataframe de exercicios escolhidos
exercicios_escolhidos = pd.DataFrame(columns=["Bloco", "PM", "Exercicio", "Series", "Reps"])

# contadores pra fazer alteracao entre emp_h e emp_v
emp_counter = 0
pux_counter = 0
aux_counter = 0


# iterar sobre dict_lista_volumes para obter o grupo muscular e a qtd de series
for grupo, lista_reps in dict_lista_volumes.items():
    # Algoritmo de escolher exercícios
    # A ideia é puxar um exercicio aleatorio para cada grupo muscular, e associar
    # a ele uma serie, cujo valor foi pre-determinado em dict_lista_volumes
    # se cair no caso emp, pux ou aux, vai alternando entre H e V para obter uma diversificacao
    # nos ex escolhidos
    # priorizar exercicios no dia
    if lista_reps:
        for serie in lista_reps:
            if grupo not in ['EMP', 'PUX', 'ESTAB', 'AUX']:
                df_filtrado = exercicios[exercicios['Id_Grupo_Muscular'] == grupo]
                if not df_filtrado.empty:
                    index_df_filtrado = df_filtrado.index
                    index_escolhido = np.random.choice(index_df_filtrado)
                    prioridade = exercicios.loc[index_escolhido, 'Prioridade no dia']
                    ex_escolhido = exercicios.loc[index_escolhido, 'Exercicio']
                    linha = {"Bloco": '-', "PM": grupo, "Exercicio": ex_escolhido, "Series": serie, "Reps": n_reps, "Prioridade": prioridade}
                    exercicios_escolhidos = exercicios_escolhidos._append(linha, ignore_index=True)
            elif grupo == 'EMP':
                if emp_counter % 2 == 0:
                    df_filtrado = exercicios[exercicios['Id_Grupo_Muscular'] == grupo + '_H']
                else:
                    df_filtrado = exercicios[exercicios['Id_Grupo_Muscular'] == grupo + '_V']
                emp_counter += 1

                if not df_filtrado.empty:
                    index_df_filtrado = df_filtrado.index
                    index_escolhido = np.random.choice(index_df_filtrado)
                    prioridade = exercicios.loc[index_escolhido, 'Prioridade no dia']
                    ex_escolhido = exercicios.loc[index_escolhido, 'Exercicio']
                    pm_grupo = grupo + '_H' if emp_counter % 2 == 0 else grupo + '_V'
                    linha = {"Bloco": '-', "PM": pm_grupo, "Exercicio": ex_escolhido, "Series": serie, "Reps": n_reps, "Prioridade": prioridade}
                    exercicios_escolhidos = exercicios_escolhidos._append(linha, ignore_index=True)
            elif grupo == 'PUX':
                if pux_counter % 2 == 0:
                    df_filtrado = exercicios[exercicios['Id_Grupo_Muscular'] == grupo + '_H']
                else:
                    df_filtrado = exercicios[exercicios['Id_Grupo_Muscular'] == grupo + '_V']
                pux_counter += 1

                if not df_filtrado.empty:
                    index_df_filtrado = df_filtrado.index
                    index_escolhido = np.random.choice(index_df_filtrado)
                    prioridade = exercicios.loc[index_escolhido, 'Prioridade no dia']
                    ex_escolhido = exercicios.loc[index_escolhido, 'Exercicio']
                    pm_grupo = grupo + '_H' if pux_counter % 2 == 0 else grupo + '_V'
                    linha = {"Bloco": '-', "PM": pm_grupo, "Exercicio": ex_escolhido, "Series": serie, "Reps": n_reps, "Prioridade": prioridade}
                    exercicios_escolhidos = exercicios_escolhidos._append(linha, ignore_index=True)
            elif grupo == 'AUX':
                if aux_counter % 2 == 0:
                    df_filtrado = exercicios[exercicios['Id_Grupo_Muscular'] == grupo + '_EMP']
                else:
                    df_filtrado = exercicios[exercicios['Id_Grupo_Muscular'] == grupo + '_PUX']
                aux_counter += 1

                if not df_filtrado.empty:
                    index_df_filtrado = df_filtrado.index
                    index_escolhido = np.random.choice(index_df_filtrado)
                    prioridade = exercicios.loc[index_escolhido, 'Prioridade no dia']
                    ex_escolhido = exercicios.loc[index_escolhido, 'Exercicio']
                    pm_grupo = grupo + '_EMP' if aux_counter % 2 == 0 else grupo + '_PUX'
                    linha = {"Bloco": '-', "PM": pm_grupo, "Exercicio": ex_escolhido, "Series": serie, "Reps": n_reps, "Prioridade": prioridade}
                    exercicios_escolhidos = exercicios_escolhidos._append(linha, ignore_index=True)

            elif grupo == 'ESTAB':
                df_filtrado = exercicios[exercicios['Id_Grupo_Muscular'] == 'EST']
                if not df_filtrado.empty:
                    index_df_filtrado = df_filtrado.index
                    index_escolhido = np.random.choice(index_df_filtrado)
                    prioridade = exercicios.loc[index_escolhido, 'Prioridade no dia']
                    ex_escolhido = exercicios.loc[index_escolhido, 'Exercicio']
                    linha = {"Bloco": '-', "PM": grupo, "Exercicio": ex_escolhido, "Series": serie, "Reps": n_reps, "Prioridade": prioridade}
                    exercicios_escolhidos = exercicios_escolhidos._append(linha, ignore_index=True)

#ordena o dataframe com 'POT' sendo o primeiro e considerando as prioridades no dia.
categoria_para_inicio = 'POT'
exercicios_escolhidos['Ordenacao'] = (exercicios_escolhidos['PM'] != categoria_para_inicio).astype(int)
exercicios_escolhidos = exercicios_escolhidos.sort_values(by=['Prioridade', 'Ordenacao'], ascending=True).drop(columns='Ordenacao').reset_index()

#logica para dividir os blocos em grupos de 2 ou de 3
if len(exercicios_escolhidos)//Frequencia > 5:
    letras_blocos = ['A','B','C']
    exercicios_escolhidos['Bloco'] = (['A'] * (((len(exercicios_escolhidos) // 3) + 1) // Frequencia) * Frequencia +
                                      ['B'] * (((len(exercicios_escolhidos) // 3) + 1) // Frequencia) * Frequencia +
                                      ['C'] * (len(exercicios_escolhidos) - 2*(((len(exercicios_escolhidos) // 3) + 1) // Frequencia) * Frequencia))
else:
    letras_blocos = ['A','B']
    exercicios_escolhidos['Bloco'] = (['A'] * (((len(exercicios_escolhidos)//2) + 1)//Frequencia)*Frequencia +
                                      ['B'] * (len(exercicios_escolhidos) - (((len(exercicios_escolhidos)//2) + 1)//Frequencia)*Frequencia))

#printar exercicios na worksheet program #n_program
i = 0
row_exercise_snp = FIRST_PROGRAM_LINE+SNP_EXERCISE_LINE
column_exercise_snp = FIRST_PROGRAM_COL+SNP_EXERCISE_1_COL
for index, row in exercicios_escolhidos.iterrows():
    output_sheet.cell(row=row_exercise_snp, column=column_exercise_snp, value=exercicios_escolhidos['Exercicio'].iloc[index])
    output_sheet.cell(row=row_exercise_snp, column=column_exercise_snp + SNP_SERIE_DELTA_COL, value=exercicios_escolhidos['Series'].iloc[index])
    output_sheet.cell(row=row_exercise_snp, column=column_exercise_snp + SNP_BLOCO_DELTA_COL, value=exercicios_escolhidos['Bloco'].iloc[index])
    output_sheet.cell(row=row_exercise_snp, column=column_exercise_snp + SNP_PM_DELTA_COL, value=exercicios_escolhidos['PM'].iloc[index])
    if exercicios_escolhidos['PM'].iloc[index] == 'POT':
        output_sheet.cell(row=row_exercise_snp, column=column_exercise_snp + SNP_REP_DELTA_COL,value='5-10')
    else: output_sheet.cell(row=row_exercise_snp, column=column_exercise_snp + SNP_REP_DELTA_COL,value=exercicios_escolhidos['Reps'].iloc[index])
    column_exercise_snp = column_exercise_snp + COLUMNS_OCCUPIED_BY_PROGRAM + COLUMNS_BETWEEN_PROGRAMS
    i += 1
    if i == Frequencia:
        column_exercise_snp = FIRST_PROGRAM_COL + SNP_EXERCISE_1_COL
        row_exercise_snp += 1
        i = 0

#inverte a ordem dos exercicios de dias pares quando o treino for fullbody
if divisao == 'FULLBODY':
    for dia in range(2,Frequencia,2):
        lista_linhas = []
        column_exercise_snp = FIRST_PROGRAM_COL+ SNP_EXERCISE_1_COL + (dia-1)*(COLUMNS_OCCUPIED_BY_PROGRAM + COLUMNS_BETWEEN_PROGRAMS)
        row_exercise_snp = FIRST_PROGRAM_LINE+SNP_EXERCISE_LINE
        exercices = []
        while(output_sheet.cell(row=row_exercise_snp,column=column_exercise_snp+SNP_PM_DELTA_COL).value != None):
            if output_sheet.cell(row=row_exercise_snp,column=column_exercise_snp+SNP_PM_DELTA_COL).value != 'POT':
                lista_linhas.append(row_exercise_snp)
                exercices.append([output_sheet.cell(row=row_exercise_snp,column=column_exercise_snp+SNP_PM_DELTA_COL).value,
                                  output_sheet.cell(row=row_exercise_snp,column=column_exercise_snp).value,
                                  output_sheet.cell(row=row_exercise_snp,column=column_exercise_snp+SNP_SERIE_DELTA_COL).value])
            row_exercise_snp = row_exercise_snp + 1
        for row in lista_linhas:
            output_sheet.cell(row=row, column=column_exercise_snp+SNP_PM_DELTA_COL,value=exercices[-1][0])
            output_sheet.cell(row=row, column=column_exercise_snp, value=exercices[-1][1])
            output_sheet.cell(row=row, column=column_exercise_snp+SNP_SERIE_DELTA_COL, value=exercices[-1][2])
            exercices.pop()

new_input_sheet = workbook.copy_worksheet(workbook["In Programa " + str(n_program)])
new_input_sheet.title = "In Programa " + str(n_program + 1)

new_input_sheet.cell(row=INPUT_HEADER_ROWS+1, column=2,value='')
new_input_sheet.cell(row=INPUT_HEADER_ROWS+2, column=2,value='')
new_input_sheet.cell(row=INPUT_HEADER_ROWS+2, column=8,value='')
new_input_sheet.cell(row=INPUT_HEADER_ROWS+3, column=2,value='')
# Getting Volumes
for i in range(7):
    new_input_sheet.cell(row=INPUT_HEADER_ROWS+5, column=i+2,value='')


workbook.save(nome_novo_treino)